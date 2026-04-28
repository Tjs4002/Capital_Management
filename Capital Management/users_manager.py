"""
Capital Management - Users Data Manager
This script manages users data in Excel format
Install required packages: pip install openpyxl
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

class UsersManager:
    def __init__(self, filename='users_data.xlsx'):
        self.filename = filename
        self.workbook = None
        self.sheet = None
        self.load_or_create_workbook()
    
    def load_or_create_workbook(self):
        """Load existing workbook or create a new one"""
        if os.path.exists(self.filename):
            self.workbook = openpyxl.load_workbook(self.filename)
            self.sheet = self.workbook.active
        else:
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = 'Users'
            self.create_headers()
            self.add_sample_data()
            self.save()
    
    def create_headers(self):
        """Create header row with formatting"""
        headers = ['ID', 'Name', 'Email', 'Phone', 'Role', 'Status', 'Joined Date']
        self.sheet.append(headers)
        
        # Format headers
        header_fill = PatternFill(start_color='052377', end_color='052377', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in self.sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        self.sheet.column_dimensions['A'].width = 5
        self.sheet.column_dimensions['B'].width = 20
        self.sheet.column_dimensions['C'].width = 25
        self.sheet.column_dimensions['D'].width = 18
        self.sheet.column_dimensions['E'].width = 12
        self.sheet.column_dimensions['F'].width = 12
        self.sheet.column_dimensions['G'].width = 15
    
    def add_sample_data(self):
        """Add sample user data"""
        sample_users = [
            [1, 'John Doe', 'john.doe@capital.com', '+1 (555) 123-4567', 'Admin', 'Active', '2026-01-15'],
            [2, 'Sarah Smith', 'sarah.smith@capital.com', '+1 (555) 234-5678', 'Manager', 'Active', '2026-02-01'],
            [3, 'Mike Johnson', 'mike.johnson@capital.com', '+1 (555) 345-6789', 'User', 'Active', '2026-02-10'],
            [4, 'Lisa Anderson', 'lisa.anderson@capital.com', '+1 (555) 456-7890', 'Manager', 'Inactive', '2026-01-20'],
            [5, 'David Wilson', 'david.wilson@capital.com', '+1 (555) 567-8901', 'User', 'Active', '2026-02-15'],
            [6, 'Emily Brown', 'emily.brown@capital.com', '+1 (555) 678-9012', 'Manager', 'Active', '2026-02-20'],
            [7, 'Robert Taylor', 'robert.taylor@capital.com', '+1 (555) 789-0123', 'User', 'Pending', '2026-03-01'],
            [8, 'Jennifer Lee', 'jennifer.lee@capital.com', '+1 (555) 890-1234', 'User', 'Active', '2026-02-25'],
            [9, 'Charles Martin', 'charles.martin@capital.com', '+1 (555) 901-2345', 'Admin', 'Active', '2026-01-10'],
            [10, 'Patricia Garcia', 'patricia.garcia@capital.com', '+1 (555) 012-3456', 'Manager', 'Active', '2026-01-25'],
        ]
        
        for user in sample_users:
            self.sheet.append(user)
        
        # Format data rows
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in self.sheet.iter_rows(min_row=2, max_row=len(sample_users)+1):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def add_user(self, name, email, phone, role, status='Active'):
        """Add a new user"""
        if not name or not email:
            raise ValueError("Name and Email are required")
        
        # Get next ID
        max_id = 0
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and isinstance(row[0], int):
                max_id = max(max_id, row[0])
        
        new_id = max_id + 1
        joined_date = datetime.now().strftime('%Y-%m-%d')
        
        self.sheet.append([new_id, name, email, phone, role, status, joined_date])
        self.save()
        return new_id
    
    def update_user(self, user_id, **kwargs):
        """Update user data"""
        for row in self.sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == user_id:
                if 'name' in kwargs:
                    row[1].value = kwargs['name']
                if 'email' in kwargs:
                    row[2].value = kwargs['email']
                if 'phone' in kwargs:
                    row[3].value = kwargs['phone']
                if 'role' in kwargs:
                    row[4].value = kwargs['role']
                if 'status' in kwargs:
                    row[5].value = kwargs['status']
                self.save()
                return True
        return False
    
    def delete_user(self, user_id):
        """Delete a user by ID"""
        for idx, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=False), start=2):
            if row[0].value == user_id:
                self.sheet.delete_rows(idx, 1)
                self.save()
                return True
        return False
    
    def get_all_users(self):
        """Get all users as list of dictionaries"""
        users = []
        headers = [cell.value for cell in self.sheet[1]]
        
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Skip empty rows
                user_dict = dict(zip(headers, row))
                users.append(user_dict)
        
        return users
    
    def get_user_stats(self):
        """Get statistics about users"""
        users = self.get_all_users()
        
        stats = {
            'total': len(users),
            'active': len([u for u in users if u['Status'] == 'Active']),
            'inactive': len([u for u in users if u['Status'] == 'Inactive']),
            'pending': len([u for u in users if u['Status'] == 'Pending']),
            'admins': len([u for u in users if u['Role'] == 'Admin']),
            'managers': len([u for u in users if u['Role'] == 'Manager']),
            'regular_users': len([u for u in users if u['Role'] == 'User']),
        }
        
        return stats
    
    def save(self):
        """Save the workbook"""
        self.workbook.save(self.filename)
        print(f"Data saved to {self.filename}")
    
    def print_all_users(self):
        """Print all users in a formatted table"""
        users = self.get_all_users()
        
        print("\n" + "="*100)
        print(f"{'ID':<5} {'Name':<20} {'Email':<25} {'Phone':<18} {'Role':<12} {'Status':<12} {'Joined':<12}")
        print("="*100)
        
        for user in users:
            print(f"{user['ID']:<5} {user['Name']:<20} {user['Email']:<25} {user['Phone']:<18} {user['Role']:<12} {user['Status']:<12} {user['Joined Date']:<12}")
        
        print("="*100)
        stats = self.get_user_stats()
        print(f"\nTotal Users: {stats['total']} | Active: {stats['active']} | Inactive: {stats['inactive']} | Pending: {stats['pending']}")
        print(f"Admins: {stats['admins']} | Managers: {stats['managers']} | Regular Users: {stats['regular_users']}\n")


def main():
    """Main function with CLI interface"""
    manager = UsersManager('users_data.xlsx')
    
    while True:
        print("\n--- Capital Management Users Manager ---")
        print("1. View all users")
        print("2. Add new user")
        print("3. Update user")
        print("4. Delete user")
        print("5. Show statistics")
        print("6. Exit")
        
        choice = input("\nEnter your choice (1-6): ").strip()
        
        if choice == '1':
            manager.print_all_users()
        
        elif choice == '2':
            name = input("Enter name: ").strip()
            email = input("Enter email: ").strip()
            phone = input("Enter phone: ").strip()
            role = input("Enter role (Admin/Manager/User): ").strip()
            status = input("Enter status (Active/Inactive/Pending) [Active]: ").strip() or 'Active'
            
            try:
                user_id = manager.add_user(name, email, phone, role, status)
                print(f"User added successfully with ID: {user_id}")
            except ValueError as e:
                print(f"Error: {e}")
        
        elif choice == '3':
            manager.print_all_users()
            user_id = int(input("Enter user ID to update: ").strip())
            
            updates = {}
            if input("Update name? (y/n): ").lower() == 'y':
                updates['name'] = input("New name: ").strip()
            if input("Update email? (y/n): ").lower() == 'y':
                updates['email'] = input("New email: ").strip()
            if input("Update phone? (y/n): ").lower() == 'y':
                updates['phone'] = input("New phone: ").strip()
            if input("Update role? (y/n): ").lower() == 'y':
                updates['role'] = input("New role: ").strip()
            if input("Update status? (y/n): ").lower() == 'y':
                updates['status'] = input("New status: ").strip()
            
            if manager.update_user(user_id, **updates):
                print("User updated successfully")
            else:
                print("User not found")
        
        elif choice == '4':
            manager.print_all_users()
            user_id = int(input("Enter user ID to delete: ").strip())
            
            if manager.delete_user(user_id):
                print("User deleted successfully")
            else:
                print("User not found")
        
        elif choice == '5':
            stats = manager.get_user_stats()
            print("\n--- User Statistics ---")
            print(f"Total Users: {stats['total']}")
            print(f"Active: {stats['active']}")
            print(f"Inactive: {stats['inactive']}")
            print(f"Pending: {stats['pending']}")
            print(f"Admins: {stats['admins']}")
            print(f"Managers: {stats['managers']}")
            print(f"Regular Users: {stats['regular_users']}")
        
        elif choice == '6':
            print("Goodbye!")
            break
        
        else:
            print("Invalid choice. Please try again.")


if __name__ == '__main__':
    main()
